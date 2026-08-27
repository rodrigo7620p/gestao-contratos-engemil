from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from contract_utils import contract_duration_months, normalize_agency_name, parse_brazilian_number
from db import connect, init_db, iso, refresh_contract_lifecycle

SKIP_SHEETS = {"BASE", "TESTE", "DASHBOARD", "ÍNDICES", "BACKLOG"}


def val(cells, address, default=None):
    value = cells.get(address)
    return default if value in (None, "") else value


def number(value, default=0.0):
    try:
        return parse_brazilian_number(value, default)
    except ValueError:
        return default


def read_cells_once(ws):
    """Read a read-only worksheet sequentially; random cell access is very slow."""
    cells = {}
    max_row = min(ws.max_row, 180)
    for row_number, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=43, values_only=True), start=1
    ):
        for column_number, value in enumerate(row, start=1):
            if value not in (None, ""):
                cells[f"{get_column_letter(column_number)}{row_number}"] = value
    return cells, max_row


def find_amendment_header(cells, max_row):
    for row in range(35, min(max_row, 80) + 1):
        if str(val(cells, f"B{row}", "")).strip().upper() == "TERMO":
            return row
    return None


def parse_contract(cells, sheet_name):
    cost_center = str(val(cells, "V4", "")).strip()
    client = str(val(cells, "B8", "")).strip()
    if not cost_center or not client:
        return None
    contract = {
        "cost_center": cost_center,
        "client": normalize_agency_name(client),
        "contract_number": str(val(cells, "AG4", "")).strip(),
        "bid_number": str(val(cells, "S8", "")).strip(),
        "uasg": str(val(cells, "Y8", "")).strip(),
        "category": str(val(cells, "M3", "")).strip(),
        "procurement_method": str(val(cells, "AC16", "")).strip(),
        "object": str(val(cells, "B11", "")).strip(),
        "signature_date": iso(val(cells, "AC8")),
        "start_date": iso(val(cells, "AH25", val(cells, "AH11"))),
        "end_date": iso(val(cells, "AH26", val(cells, "AH12"))),
        "original_value": number(val(cells, "AC19")),
        "current_value": number(val(cells, "AC22", val(cells, "AC19"))),
        "status": "ATIVO",
        "manager_name": str(val(cells, "B25", "")).strip(),
        "manager_email": str(val(cells, "P25", "")).strip(),
        "observations": str(val(cells, "B31", "")).strip(),
        "source_sheet": sheet_name,
    }
    for instrument in parse_amendments(cells, 180, include_initial=True):
        if is_initial_contract_instrument(instrument):
            contract["original_start_date"] = (
                instrument.get("start_date") or contract["start_date"]
            )
            contract["original_end_date"] = (
                instrument.get("end_date") or contract["end_date"]
            )
            break
    return contract


def is_initial_contract_instrument(instrument):
    ordinal = str(instrument.get("ordinal") or "").strip().upper()
    kind = str(instrument.get("kind") or "").strip().upper()
    return (
        ordinal in {"INICIAL", "CONTRATO INICIAL"}
        and kind in {"CONTRATO", "CONTRATO INICIAL"}
    )


def parse_amendments(cells, max_row, include_initial=False):
    header = find_amendment_header(cells, max_row)
    if not header:
        return []
    result = []
    for row in range(header + 1, max_row + 1):
        ordinal, kind = val(cells, f"B{row}"), val(cells, f"E{row}")
        if ordinal in (None, "") and kind in (None, ""):
            continue
        start_date = iso(val(cells, f"V{row}"))
        end_date = iso(val(cells, f"AA{row}"))
        imported_duration = int(number(val(cells, f"S{row}")))
        instrument = {
                "ordinal": str(ordinal or "").strip(),
                "kind": str(kind or "").strip(),
                "value": number(val(cells, f"K{row}")),
                "duration_months": imported_duration or contract_duration_months(
                    start_date, end_date
                ),
                "start_date": start_date,
                "end_date": end_date,
                "guarantee_status": str(val(cells, f"AF{row}", "")).strip(),
                "art_status": str(val(cells, f"AQ{row}", "")).strip(),
            }
        if include_initial or not is_initial_contract_instrument(instrument):
            result.append(instrument)
    return result


def import_workbook(path: str | Path, replace_amendments: bool = False):
    init_db()
    workbook = load_workbook(path, data_only=True, read_only=True)
    counters = {"contracts": 0, "amendments": 0, "skipped": 0}
    processed_contract_ids = []
    with connect() as conn:
        for sheet_name in workbook.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            ws = workbook[sheet_name]
            cells, max_row = read_cells_once(ws)
            contract = parse_contract(cells, sheet_name)
            if not contract:
                counters["skipped"] += 1
                continue
            columns = ",".join(contract)
            placeholders = ",".join("?" for _ in contract)
            updates = ",".join(f"{k}=excluded.{k}" for k in contract if k != "cost_center")
            conn.execute(
                f"INSERT INTO contracts({columns}) VALUES({placeholders}) "
                f"ON CONFLICT(cost_center) DO UPDATE SET {updates},updated_at=CURRENT_TIMESTAMP",
                tuple(contract.values()),
            )
            contract_id = conn.execute(
                "SELECT id FROM contracts WHERE cost_center=?", (contract["cost_center"],)
            ).fetchone()["id"]
            processed_contract_ids.append(contract_id)
            counters["contracts"] += 1
            if replace_amendments:
                conn.execute("DELETE FROM amendments WHERE contract_id=?", (contract_id,))
            existing = conn.execute(
                "SELECT COUNT(*) AS n FROM amendments WHERE contract_id=?", (contract_id,)
            ).fetchone()["n"]
            if not existing:
                for amendment in parse_amendments(cells, max_row):
                    keys = ["contract_id", *amendment.keys()]
                    conn.execute(
                        f"INSERT INTO amendments({','.join(keys)}) VALUES({','.join('?' for _ in keys)})",
                        (contract_id, *amendment.values()),
                    )
                    counters["amendments"] += 1
    for contract_id in processed_contract_ids:
        refresh_contract_lifecycle(contract_id)
    return counters


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Importa a planilha de contratos para o SQLite.")
    parser.add_argument("arquivo")
    parser.add_argument("--substituir-aditivos", action="store_true")
    args = parser.parse_args()
    print(import_workbook(args.arquivo, args.substituir_aditivos))
