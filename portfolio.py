from __future__ import annotations

from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from contract_utils import today_brt


def parse_date(value):
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def annual_allocation(start_value, end_value, total_value, year: int) -> float:
    start, end = parse_date(start_value), parse_date(end_value)
    if not start or not end or end < start:
        return 0.0
    period_start, period_end = date(year, 1, 1), date(year, 12, 31)
    overlap_start, overlap_end = max(start, period_start), min(end, period_end)
    if overlap_end < overlap_start:
        return 0.0
    total_days = (end - start).days + 1
    overlap_days = (overlap_end - overlap_start).days + 1
    return round(float(total_value or 0) * overlap_days / total_days, 2)


def remaining_value(start_value, end_value, total_value, as_of: date | None = None) -> float:
    start, end = parse_date(start_value), parse_date(end_value)
    as_of = as_of or today_brt()
    if not start or not end or end < start:
        return 0.0
    if as_of <= start:
        return float(total_value or 0)
    if as_of > end:
        return 0.0
    total_days = (end - start).days + 1
    remaining_days = (end - as_of).days + 1
    return round(float(total_value or 0) * remaining_days / total_days, 2)


def backlog_rows(contracts, start_year: int | None = None, years: int = 6):
    start_year = start_year or today_brt().year
    result = []
    for item, contract in enumerate(contracts, start=1):
        row = {
            "Item": item,
            "Centro de custo": contract["cost_center"],
            "Contratante": contract["client"],
            "Contrato": contract["contract_number"],
            "Início": contract["start_date"],
            "Fim": contract["end_date"],
            "Valor atual": float(contract["current_value"] or 0),
            "Instrumento vigente": contract.get("current_instrument") or "CONTRATO",
            "Status": contract["status"],
            "Modalidade": contract["category"],
            "Responsável": contract["manager_name"],
        }
        for year in range(start_year, start_year + years):
            row[str(year)] = annual_allocation(
                contract["start_date"], contract["end_date"], contract["current_value"], year
            )
        row["Remanescente total"] = remaining_value(
            contract["start_date"], contract["end_date"], contract["current_value"]
        )
        result.append(row)
    return result


def dataframe_sheet(workbook, title: str, dataframe: pd.DataFrame):
    ws = workbook.create_sheet(title[:31])
    headers = list(dataframe.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.alignment = Alignment(horizontal="center")
    for row in dataframe.itertuples(index=False, name=None):
        ws.append(list(row))
    for index, column in enumerate(headers, start=1):
        width = min(max(len(str(column)) + 2, 12), 45)
        for cell in ws[get_column_letter(index)][1:]:
            if isinstance(cell.value, str):
                width = min(max(width, len(cell.value[:45]) + 2), 45)
            if isinstance(cell.value, (float, int)) and (
                "valor" in column.lower() or column.isdigit() or "remanescente" in column.lower()
            ):
                cell.number_format = '"R$" #,##0.00'
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def workbook_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, dataframe in sheets.items():
        dataframe_sheet(workbook, title, dataframe)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
